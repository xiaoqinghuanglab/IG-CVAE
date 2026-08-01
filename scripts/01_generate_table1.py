#!/usr/bin/env python3
"""Create manuscript Table 1 from prepared ADNI and AdNeuroMed metadata.

Source files are read-only. ADNI MRI paths are resolved relative to the project
root, so the script is portable and does not depend on stale absolute paths.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import sys
from pathlib import Path

try:
    import numpy as np
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from scipy.stats import chi2_contingency, ttest_ind
except ImportError as exc:
    raise SystemExit(
        f"Missing package {exc.name!r}. Required: numpy, pandas, scipy, openpyxl."
    ) from exc


TITLE = "Table 1. Characteristics of the study sample."
HEADER = ["Variable", "AD", "Control", "Total", "P-Value"]

# SHA256 of every displayed cell in the supplied reference table.
REFERENCE_SHA256 = (
    "076b23df739a4b1cd64800ffe4caa33fe12aa23bf71ee743bfebaace3cea80ee"
)


def normalize_name(value: str) -> str:
    return "".join(
        character.lower() for character in value if character.isalnum()
    )


def find_column(
    frame: pd.DataFrame,
    *names: str,
    required: bool = True,
) -> str | None:
    lookup = {
        normalize_name(column_name): column_name
        for column_name in frame.columns
    }

    for name in names:
        normalized = normalize_name(name)
        if normalized in lookup:
            return lookup[normalized]

    if required:
        raise KeyError(
            f"Missing column {names}; available columns: {list(frame.columns)}"
        )
    return None


def column(
    frame: pd.DataFrame,
    *names: str,
    required: bool = True,
) -> pd.Series:
    match = find_column(frame, *names, required=required)

    if match is None:
        return pd.Series(pd.NA, index=frame.index, dtype="string")

    return frame[match]




def normalize_apoe4(value: object) -> float:
    """Convert APOE allele counts or genotypes into APOE4 counts."""

    if pd.isna(value):
        return np.nan

    text = str(value).strip().upper()

    if text in {
        "",
        "NA",
        "N/A",
        "NAN",
        "NONE",
        "MISSING",
        "-9",
        "-99",
    }:
        return np.nan

    try:
        numeric_value = float(text)
    except ValueError:
        numeric_value = math.nan

    # Already encoded as number of APOE4 alleles.
    if (
        math.isfinite(numeric_value)
        and numeric_value in {0.0, 1.0, 2.0}
    ):
        return numeric_value

    # Handles forms such as 33, 34, 44, 3/4,
    # E3/E4, APOE3/APOE4, and 3.0/4.0.
    alleles = re.findall(r"[234]", text)

    if len(alleles) == 2:
        return float(
            sum(allele == "4" for allele in alleles)
        )

    raise ValueError(
        f"Unrecognized APOE genotype or allele count: {value!r}"
    )


def standardize(path: Path) -> pd.DataFrame:
    if not path.is_file():
        raise FileNotFoundError(f"Required metadata file not found: {path}")

    raw = pd.read_csv(path, dtype="string", keep_default_na=True)
    result = pd.DataFrame(index=raw.index)

    result["subject_id"] = (
        column(raw, "subject_id")
        .astype("string")
        .str.strip()
    )

    result["bids_id"] = (
        column(
            raw,
            "bids_id",
            "participant_id",
            required=False,
        )
        .astype("string")
        .str.strip()
    )

    result["diagnosis"] = (
        column(raw, "diagnosis")
        .astype("string")
        .str.strip()
        .str.upper()
        .replace(
            {
                "CONTROL": "CN",
                "CTL": "CN",
            }
        )
    )

    result["age"] = pd.to_numeric(
        column(raw, "age"),
        errors="coerce",
    )

    result["sex"] = (
        column(raw, "sex")
        .astype("string")
        .str.strip()
        .str.upper()
        .replace(
            {
                "FEMALE": "F",
                "MALE": "M",
            }
        )
    )

    result["apoe4"] = column(
        raw,
        "apoe4_allele_count",
        "apoe4",
        "apoe",
        required=False,
    ).map(normalize_apoe4)

    result["batch"] = (
        column(raw, "batch", required=False)
        .astype("string")
        .str.strip()
        .str.upper()
        .str.replace(" ", "", regex=False)
    )

    if result["subject_id"].isna().any():
        raise ValueError(f"Missing subject ID in {path}")

    if result["subject_id"].duplicated().any():
        raise ValueError(f"Duplicate subject ID in {path}")

    unexpected = sorted(
        set(result["diagnosis"].dropna()) - {"AD", "CN"}
    )

    if unexpected:
        raise ValueError(
            f"Unexpected diagnoses in {path}: {unexpected}"
        )

    return result


def load_samples(
    project_root: Path,
) -> list[tuple[str, pd.DataFrame]]:
    data_root = project_root / "data"
    adni_root = data_root / "ADNI"
    anm_root = data_root / "AdNeuroMed"

    adni_ge = standardize(
        adni_root / "Paired_data_metadata.csv"
    )

    if adni_ge["bids_id"].isna().any():
        raise ValueError(
            "ADNI paired metadata contains missing bids_id values"
        )

    # Resolve ADNI paths portably instead of trusting stale mri_abs_path.
    corrected_paths = adni_ge["bids_id"].map(
        lambda bids_id:
        adni_root / "paired_MRI" / f"{bids_id}.nii.gz"
    )

    adni_paired = adni_ge.loc[
        corrected_paths.map(Path.is_file)
    ].copy()

    samples = [
        (
            "ADNI Gene Expression sample",
            adni_ge,
        ),
        (
            "ADNI paired MRI-GE sample",
            adni_paired,
        ),
        (
            "ADNI MRI-only sample",
            standardize(
                adni_root / "mri_only_metadata.csv"
            ),
        ),
        (
            "AdNeuroMed Gene Expression sample",
            standardize(
                anm_root / "gene_expression_metadata.csv"
            ),
        ),
        (
            "AdNeuroMed paired MRI-GE sample",
            standardize(
                anm_root / "Paired_data_metadata.csv"
            ),
        ),
        (
            "AdNeuroMed MRI-only sample",
            standardize(
                anm_root / "mri_only_metadata.csv"
            ),
        ),
    ]

    expected_counts = {
        "ADNI Gene Expression sample": (90, 212, 302),
        "ADNI paired MRI-GE sample": (85, 204, 289),
        "ADNI MRI-only sample": (546, 660, 1206),
        "AdNeuroMed Gene Expression sample": (181, 169, 350),
        "AdNeuroMed paired MRI-GE sample": (46, 52, 98),
        "AdNeuroMed MRI-only sample": (10, 3, 13),
    }

    for label, frame in samples:
        observed = (
            int((frame["diagnosis"] == "AD").sum()),
            int((frame["diagnosis"] == "CN").sum()),
            len(frame),
        )

        if observed != expected_counts[label]:
            raise ValueError(
                f"Unexpected cohort count for {label}: "
                f"{observed}; expected {expected_counts[label]}"
            )

    return samples


def mean_sd(values: pd.Series) -> str:
    values = (
        pd.to_numeric(values, errors="coerce")
        .dropna()
        .astype(float)
    )

    return (
        f"{values.mean():.2f} ± "
        f"{values.std(ddof=1):.2f}"
    )


def count_percent(count: int, denominator: int) -> str:
    percentage = 100.0 * count / denominator
    return f"{percentage:.2f}% ({count})"


def format_p(value: float) -> str:
    if not math.isfinite(value):
        return ""

    if value < 1e-4:
        return f"{value:.2E}"

    return (
        f"{value:.3f}".rstrip("0").rstrip(".")
        or "0"
    )


def welch_p(
    ad: pd.Series,
    control: pd.Series,
) -> float:
    ad = (
        pd.to_numeric(ad, errors="coerce")
        .dropna()
        .astype(float)
    )

    control = (
        pd.to_numeric(control, errors="coerce")
        .dropna()
        .astype(float)
    )

    result = ttest_ind(
        ad,
        control,
        equal_var=False,
        nan_policy="omit",
        alternative="two-sided",
    )

    return float(result.pvalue)


def chi_square_p(
    ad: pd.Series,
    control: pd.Series,
    categories: list[object],
) -> float:
    observed = np.asarray(
        [
            [
                int((ad == category).sum())
                for category in categories
            ],
            [
                int((control == category).sum())
                for category in categories
            ],
        ],
        dtype=int,
    )

    # Remove categories absent from both groups.
    observed = observed[:, observed.sum(axis=0) > 0]

    if observed.shape[1] < 2:
        return math.nan

    # correction=True applies Yates correction to 2x2 tables.
    return float(
        chi2_contingency(
            observed,
            correction=True,
        ).pvalue
    )


def build_section(
    label: str,
    frame: pd.DataFrame,
) -> list[list[str]]:
    ad = frame.loc[frame["diagnosis"] == "AD"]
    control = frame.loc[frame["diagnosis"] == "CN"]
    total = frame

    rows = [
        [label, "", "", "", ""],
        [
            "",
            f"N={len(ad)}",
            f"N={len(control)}",
            f"N={len(total)}",
            "",
        ],
        [
            "Age (Mean ± SD)",
            mean_sd(ad["age"]),
            mean_sd(control["age"]),
            mean_sd(total["age"]),
            format_p(
                welch_p(
                    ad["age"],
                    control["age"],
                )
            ),
        ],
    ]

    sex_p = chi_square_p(
        ad["sex"],
        control["sex"],
        ["F", "M"],
    )

    for index, category in enumerate(["F", "M"]):
        rows.append(
            [
                f"Sex: {category}",
                count_percent(
                    int((ad["sex"] == category).sum()),
                    len(ad),
                ),
                count_percent(
                    int((control["sex"] == category).sum()),
                    len(control),
                ),
                count_percent(
                    int((total["sex"] == category).sum()),
                    len(total),
                ),
                format_p(sex_p) if index == 0 else "",
            ]
        )

    apoe_categories = [0.0, 1.0, 2.0]

    apoe_p = chi_square_p(
        ad["apoe4"],
        control["apoe4"],
        apoe_categories,
    )

    ad_apoe_denominator = int(
        ad["apoe4"].notna().sum()
    )

    control_apoe_denominator = int(
        control["apoe4"].notna().sum()
    )

    total_apoe_denominator = int(
        total["apoe4"].notna().sum()
    )

    for index, category in enumerate(apoe_categories):
        rows.append(
            [
                f"APOE4: {int(category)}",
                count_percent(
                    int((ad["apoe4"] == category).sum()),
                    ad_apoe_denominator,
                ),
                count_percent(
                    int((control["apoe4"] == category).sum()),
                    control_apoe_denominator,
                ),
                count_percent(
                    int((total["apoe4"] == category).sum()),
                    total_apoe_denominator,
                ),
                format_p(apoe_p) if index == 0 else "",
            ]
        )

    if total["apoe4"].isna().any():
        rows.append(
            [
                "APOE4: Missing",
                count_percent(
                    int(ad["apoe4"].isna().sum()),
                    len(ad),
                ),
                count_percent(
                    int(control["apoe4"].isna().sum()),
                    len(control),
                ),
                count_percent(
                    int(total["apoe4"].isna().sum()),
                    len(total),
                ),
                "",
            ]
        )

    batches = sorted(
        {
            str(value)
            for value in total["batch"].dropna()
        },
        key=lambda value: (
            re.sub(r"\d+$", "", value),
            (
                int(re.search(r"\d+$", value).group())
                if re.search(r"\d+$", value)
                else -1
            ),
        ),
    )

    if batches:
        batch_p = chi_square_p(
            ad["batch"],
            control["batch"],
            batches,
        )

        for index, batch in enumerate(batches):
            rows.append(
                [
                    f"Batch: {batch}",
                    count_percent(
                        int((ad["batch"] == batch).sum()),
                        len(ad),
                    ),
                    count_percent(
                        int((control["batch"] == batch).sum()),
                        len(control),
                    ),
                    count_percent(
                        int((total["batch"] == batch).sum()),
                        len(total),
                    ),
                    format_p(batch_p) if index == 0 else "",
                ]
            )

    return rows


def build_table(
    samples: list[tuple[str, pd.DataFrame]],
) -> list[list[str]]:
    rows: list[list[str]] = []

    for label, frame in samples:
        rows.extend(
            build_section(label, frame)
        )

    return rows


def table_digest(rows: list[list[str]]) -> str:
    payload = json.dumps(
        rows,
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")

    return hashlib.sha256(payload).hexdigest()


def write_csv(
    path: Path,
    rows: list[list[str]],
) -> None:
    with path.open(
        "w",
        encoding="utf-8",
        newline="",
    ) as handle:
        writer = csv.writer(handle)

        writer.writerow(
            [TITLE, "", "", "", ""]
        )

        writer.writerow(
            ["", "", "", "", ""]
        )

        writer.writerow(HEADER)
        writer.writerows(rows)


def write_excel(
    path: Path,
    rows: list[list[str]],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Table 1"
    sheet.sheet_view.showGridLines = False

    sheet.append([TITLE, "", "", "", ""])
    sheet.append(["", "", "", "", ""])
    sheet.append(HEADER)

    for row in rows:
        sheet.append(row)

    # Store p-values as numeric cells while retaining
    # the reference display format.
    for row_number in range(4, sheet.max_row + 1):
        cell = sheet.cell(row_number, 5)

        if isinstance(cell.value, str) and cell.value:
            displayed = cell.value
            cell.value = float(displayed)

            if "E" in displayed:
                cell.number_format = "0.00E+00"
            elif displayed in {"0", "1"}:
                cell.number_format = "0"
            else:
                cell.number_format = "0.###"

    sheet.merge_cells("A1:E1")
    sheet.freeze_panes = "A4"

    thin = Side(
        style="thin",
        color="000000",
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    header_fill = PatternFill(
        "solid",
        fgColor="D9E1F2",
    )

    section_fill = PatternFill(
        "solid",
        fgColor="E7E6E6",
    )

    section_names = {
        row[0]
        for row in rows
        if row[0].endswith("sample")
    }

    for row in sheet.iter_rows(
        min_row=1,
        max_row=sheet.max_row,
        min_col=1,
        max_col=5,
    ):
        for cell in row:
            cell.font = Font(
                name="Calibri",
                size=9,
            )

            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if cell.column == 1
                    else "right"
                ),
                vertical="center",
            )

            if cell.row != 2:
                cell.border = border

    sheet["A1"].font = Font(
        name="Calibri",
        size=10,
        bold=True,
    )

    for cell in sheet[3]:
        cell.font = Font(
            name="Calibri",
            size=9,
            bold=True,
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

    for row_number in range(
        4,
        sheet.max_row + 1,
    ):
        if (
            sheet.cell(row_number, 1).value
            in section_names
        ):
            for column_number in range(1, 6):
                cell = sheet.cell(
                    row_number,
                    column_number,
                )

                cell.fill = section_fill

                cell.font = Font(
                    name="Calibri",
                    size=9,
                    bold=True,
                )

    column_widths = {
        "A": 31,
        "B": 21,
        "C": 21,
        "D": 21,
        "E": 13,
    }

    for name, width in column_widths.items():
        sheet.column_dimensions[name].width = width

    sheet.row_dimensions[1].height = 18
    sheet.row_dimensions[2].height = 7
    sheet.row_dimensions[3].height = 17

    for row_number in range(
        4,
        sheet.max_row + 1,
    ):
        sheet.row_dimensions[row_number].height = 15

    sheet.print_area = f"A1:E{sheet.max_row}"
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    workbook.save(path)


def print_table(
    rows: list[list[str]],
) -> None:
    grid = [HEADER, *rows]

    widths = [
        max(
            len(str(row[index]))
            for row in grid
        )
        for index in range(5)
    ]

    def line(character: str) -> str:
        return (
            "+"
            + "+".join(
                character * (width + 2)
                for width in widths
            )
            + "+"
        )

    def display(row: list[str]) -> str:
        cells = []

        for index, value in enumerate(row):
            if index == 0:
                cells.append(
                    f" {value:<{widths[index]}} "
                )
            else:
                cells.append(
                    f" {value:>{widths[index]}} "
                )

        return "|" + "|".join(cells) + "|"

    print(f"\n{TITLE}")
    print(line("="))
    print(display(HEADER))
    print(line("="))

    for row in rows:
        if row[0].endswith("sample"):
            print(line("-"))

        print(display(row))

    print(line("="))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=__doc__,
    )

    parser.add_argument(
        "--project-root",
        type=Path,
        default=Path(__file__).resolve().parents[1],
        help=(
            "Project root; defaults to the parent "
            "of scripts/."
        ),
    )

    parser.add_argument(
        "--strict-reference",
        action="store_true",
        help=(
            "Return a nonzero exit status if Table 1 "
            "differs from the reference."
        ),
    )

    return parser.parse_args()


def main() -> int:
    args = parse_args()

    project_root = (
        args.project_root
        .expanduser()
        .resolve()
    )

    output_directory = (
        project_root
        / "Output"
        / "Table"
    )

    output_directory.mkdir(
        parents=True,
        exist_ok=True,
    )

    samples = load_samples(project_root)
    rows = build_table(samples)

    generated_digest = table_digest(rows)

    reference_pass = (
        generated_digest
        == REFERENCE_SHA256
    )

    xlsx_path = output_directory / "Table1.xlsx"
    csv_path = output_directory / "Table1.csv"
    qc_path = output_directory / "Table1_QC.txt"

    write_excel(xlsx_path, rows)
    write_csv(csv_path, rows)

    qc_lines = [
        "Table 1 QC",
        f"Project root: {project_root}",
        (
            "ADNI MRI paths: resolved relative to "
            "data/ADNI (source CSV unchanged)"
        ),
        "Age test: two-sided Welch t-test",
        (
            "Categorical tests: Pearson chi-square; "
            "Yates correction for 2x2 tables"
        ),
        (
            "Missing APOE4: displayed but excluded "
            "from hypothesis test"
        ),
        (
            "Reference screenshot match: "
            f"{'PASS' if reference_pass else 'FAIL'}"
        ),
        f"Generated SHA256: {generated_digest}",
        f"Expected SHA256:  {REFERENCE_SHA256}",
    ]

    qc_path.write_text(
        "\n".join(qc_lines) + "\n",
        encoding="utf-8",
    )

    print_table(rows)

    print("\nCohort verification:")

    for label, frame in samples:
        ad_count = int(
            (frame["diagnosis"] == "AD").sum()
        )

        control_count = int(
            (frame["diagnosis"] == "CN").sum()
        )

        print(
            f"{label}: "
            f"AD={ad_count}, "
            f"Control={control_count}, "
            f"Total={len(frame)}"
        )

    print(f"\nCreated: {xlsx_path}")
    print(f"Created: {csv_path}")
    print(f"Created: {qc_path}")

    print(
        "Reference screenshot match: "
        f"{'PASS' if reference_pass else 'FAIL'}"
    )

    if args.strict_reference and not reference_pass:
        return 2

    return 0


if __name__ == "__main__":
    sys.exit(main())
