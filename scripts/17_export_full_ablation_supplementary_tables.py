#!/usr/bin/env python3

from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_ROOT = ROOT / "Output"
SUPPLEMENTARY_DIR = OUTPUT_ROOT / "Table" / "Supplementary"
SUPPLEMENTARY_DIR.mkdir(parents=True, exist_ok=True)

MAX_EXCEL_DATA_ROWS = 1_048_000

TABLE_CONFIG = {
    "gene": {
        "number": 5,
        "title": "Full Gene-Branch Ablation and Permutation-Importance Results",
        "filename": "Supplementary_Table_5_Gene_ablation_full.xlsx",
        "preferred_filename": "gene_permutation_importance_summary.csv",
        "source_directories": [
            OUTPUT_ROOT / "Table" / "Gene_ablation" / "dual_cohort_top500_v1",
        ],
    },
    "mri": {
        "number": 6,
        "title": "Full MRI Regional Zero-Occlusion Results",
        "filename": "Supplementary_Table_6_MRI_ablation_full.xlsx",
        "preferred_filename": "region_occlusion_summary.csv",
        "source_directories": [
            OUTPUT_ROOT / "Table" / "MRI_ablation" / "dual_cohort_independent_v1",
        ],
    },
    "cvae": {
        "number": 7,
        "title": "Full XIG-CVAE Region–Gene-Feature Ablation Results",
        "filename": "Supplementary_Table_7_CVAE_ablation_full.xlsx",
        "preferred_filename": "gene_region_projected_effects.csv",
        "source_directories": [
            OUTPUT_ROOT / "Table" / "CVAE_ablation" / "dual_cohort_independent_v1",
        ],
    },
}


def normalize_name(value: object) -> str:
    """Normalize a name for flexible matching."""
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def identify_cohort(value: object) -> Optional[str]:
    """Map common cohort names and abbreviations to manuscript labels."""
    text = str(value).strip().lower()

    if (
        "addneuromed" in text
        or "adneuro" in text
        or re.search(r"(^|[^a-z0-9])anm([^a-z0-9]|$)", text)
    ):
        return "AddNeuroMed"

    if "adni" in text:
        return "ADNI"

    return None


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Remove empty rows, empty columns, and accidental index columns."""
    if df is None or df.empty:
        return pd.DataFrame()

    df = df.copy()
    df = df.dropna(axis=0, how="all")
    df = df.dropna(axis=1, how="all")

    removable = []
    for column in df.columns:
        normalized = normalize_name(column)
        if normalized.startswith("unnamed"):
            series = df[column]
            if series.isna().all():
                removable.append(column)
            elif pd.api.types.is_numeric_dtype(series):
                expected = pd.Series(range(len(series)), index=series.index)
                if series.reset_index(drop=True).equals(expected):
                    removable.append(column)

    if removable:
        df = df.drop(columns=removable)

    return df.reset_index(drop=True)


def read_data_file(path: Path) -> List[Tuple[str, pd.DataFrame]]:
    """Read every tabular object contained in a source file."""
    suffix = path.suffix.lower()
    outputs: List[Tuple[str, pd.DataFrame]] = []

    try:
        if suffix == ".csv":
            df = pd.read_csv(path, low_memory=False)
            outputs.append(("CSV", clean_dataframe(df)))

        elif suffix == ".tsv":
            df = pd.read_csv(path, sep="\t", low_memory=False)
            outputs.append(("TSV", clean_dataframe(df)))

        elif suffix in {".xlsx", ".xls"}:
            sheets: Dict[str, pd.DataFrame] = pd.read_excel(
                path,
                sheet_name=None,
            )
            for sheet_name, df in sheets.items():
                outputs.append((sheet_name, clean_dataframe(df)))

        elif suffix == ".parquet":
            df = pd.read_parquet(path)
            outputs.append(("Parquet", clean_dataframe(df)))

    except Exception as exc:
        print(f"WARNING: Could not read {path}: {exc}", file=sys.stderr)

    return [(sheet, df) for sheet, df in outputs if not df.empty]


def find_cohort_column(df: pd.DataFrame) -> Optional[str]:
    candidates = {
        "cohort",
        "dataset",
        "study",
        "sourcecohort",
        "cohortname",
        "datasetname",
    }

    for column in df.columns:
        if normalize_name(column) in candidates:
            return column

    return None


def dataframe_relevance(stage: str, df: pd.DataFrame) -> int:
    """Require columns that indicate a real full explainability result."""
    columns = [normalize_name(c) for c in df.columns]
    joined = " ".join(columns)

    has_gene = any(
        token in joined
        for token in [
            "gene",
            "genesymbol",
            "featuregene",
            "inputfeature",
            "featurename",
        ]
    )

    has_region = any(
        token in joined
        for token in [
            "region",
            "regionname",
            "atlaslabel",
            "labelname",
            "dkregion",
            "roi",
        ]
    )

    has_importance = any(
        token in joined
        for token in [
            "importance",
            "delta",
            "auc",
            "rocauc",
            "prauc",
            "ablation",
            "permutation",
            "perturbation",
            "occlusion",
            "drop",
            "change",
            "shift",
            "similarity",
            "attribution",
            "score",
        ]
    )

    if stage == "gene":
        return int(has_gene) + int(has_importance)

    if stage == "mri":
        return int(has_region) + int(has_importance)

    if stage == "cvae":
        return int(has_gene) + int(has_region) + int(has_importance)

    return 0


def candidate_files(config: dict) -> List[Path]:
    """Collect tabular files from expected source directories."""
    supported = {".csv", ".tsv", ".xlsx", ".xls", ".parquet"}
    files: List[Path] = []

    for directory in config["source_directories"]:
        if directory.exists():
            for path in directory.rglob("*"):
                if (
                    path.is_file()
                    and path.suffix.lower() in supported
                    and not path.name.startswith("~$")
                    and "supplementary_table" not in path.name.lower()
                ):
                    files.append(path)

    files = sorted(set(files))

    preferred_filename = config.get("preferred_filename")
    if preferred_filename:
        preferred_files = [
            path for path in files
            if path.name == preferred_filename
        ]

        if not preferred_files:
            raise FileNotFoundError(
                f"Preferred source file not found: {preferred_filename}"
            )

        files = preferred_files

    return files


def split_by_cohort(
    df: pd.DataFrame,
    source_text: str,
) -> Dict[str, pd.DataFrame]:
    """Split a combined table by cohort, or infer cohort from its path."""
    output: Dict[str, pd.DataFrame] = {}
    cohort_column = find_cohort_column(df)

    if cohort_column is not None:
        mapped = df[cohort_column].map(identify_cohort)

        for cohort in ["ADNI", "AddNeuroMed"]:
            subset = df.loc[mapped == cohort].copy()
            if not subset.empty:
                output[cohort] = subset.reset_index(drop=True)

    if output:
        return output

    inferred = identify_cohort(source_text)
    if inferred is not None:
        output[inferred] = df.reset_index(drop=True)

    return output


def collect_candidates(stage: str, config: dict) -> Dict[str, List[dict]]:
    """Read candidate files and retain plausible full result tables."""
    collected: Dict[str, List[dict]] = {
        "ADNI": [],
        "AddNeuroMed": [],
    }

    files = candidate_files(config)

    print(f"\n[{stage.upper()}] Found {len(files)} candidate data files.")

    for path in files:
        relative_path = path.relative_to(ROOT)

        for sheet_name, dataframe in read_data_file(path):
            relevance = dataframe_relevance(stage, dataframe)

            required_relevance = 3 if stage == "cvae" else 2
            if relevance < required_relevance:
                continue

            source_text = f"{relative_path} {sheet_name}"
            cohort_tables = split_by_cohort(dataframe, source_text)

            for cohort, cohort_df in cohort_tables.items():
                collected[cohort].append(
                    {
                        "dataframe": cohort_df,
                        "source_file": str(relative_path),
                        "source_sheet": sheet_name,
                        "rows": len(cohort_df),
                        "columns": len(cohort_df.columns),
                        "relevance": relevance,
                    }
                )

    return collected


def select_full_table(candidates: List[dict]) -> dict:
    """
    Select the most complete relevant result table.

    Priority:
      1. largest number of rows;
      2. highest relevance;
      3. largest number of columns.
    """
    if not candidates:
        raise RuntimeError("No suitable full result table was identified.")

    return sorted(
        candidates,
        key=lambda item: (
            item["rows"],
            item["relevance"],
            item["columns"],
        ),
        reverse=True,
    )[0]


def safe_sheet_name(name: str) -> str:
    name = re.sub(r"[\[\]:*?/\\]", "_", name)
    return name[:31]


def dataframe_parts(
    dataframe: pd.DataFrame,
    cohort: str,
) -> Iterable[Tuple[str, pd.DataFrame]]:
    """Split very large tables so they remain valid Excel worksheets."""
    if len(dataframe) <= MAX_EXCEL_DATA_ROWS:
        yield cohort, dataframe
        return

    part = 1
    for start in range(0, len(dataframe), MAX_EXCEL_DATA_ROWS):
        stop = min(start + MAX_EXCEL_DATA_ROWS, len(dataframe))
        yield f"{cohort}_{part}", dataframe.iloc[start:stop].copy()
        part += 1


def format_workbook(
    workbook_path: Path,
    table_title: str,
) -> None:
    wb = load_workbook(workbook_path)

    dark_fill = PatternFill("solid", fgColor="1F4E78")
    secondary_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=14, bold=True, color="FFFFFF")
    subtitle_font = Font(size=10, italic=True, color="404040")
    header_font = Font(size=10, bold=True, color="FFFFFF")
    thin_gray = Side(style="thin", color="D9E1F2")
    border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

        if ws.title == "README":
            ws.freeze_panes = "A2"

            for cell in ws[1]:
                cell.fill = dark_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center")

            ws.column_dimensions["A"].width = 27
            ws.column_dimensions["B"].width = 90

            for row in ws.iter_rows(min_row=2):
                row[0].font = Font(bold=True)
                row[0].fill = secondary_fill
                for cell in row:
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True,
                    )
                    cell.border = border

            continue

        last_column = max(ws.max_column, 1)
        last_column_letter = get_column_letter(last_column)

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=last_column,
        )
        ws.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=last_column,
        )

        ws["A1"] = table_title
        ws["A1"].fill = dark_fill
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        source_note = ws["A2"].value or ""
        ws["A2"] = source_note
        ws["A2"].fill = secondary_fill
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 30

        header_row = 4

        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = (
            f"A{header_row}:{last_column_letter}{ws.max_row}"
        )

        sample_end = min(ws.max_row, 204)

        for column_index in range(1, last_column + 1):
            letter = get_column_letter(column_index)
            maximum = 0

            for row_index in range(4, sample_end + 1):
                value = ws.cell(row=row_index, column=column_index).value
                if value is not None:
                    maximum = max(maximum, len(str(value)))

            ws.column_dimensions[letter].width = min(max(maximum + 2, 11), 38)

        for row in ws.iter_rows(
            min_row=5,
            max_row=min(ws.max_row, 1004),
        ):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=False,
                )
                cell.border = border

    if "README" in wb.sheetnames:
        wb.active = wb.sheetnames.index("README")

    wb.save(workbook_path)


def export_stage(stage: str, config: dict) -> Path:
    candidates = collect_candidates(stage, config)

    selected: Dict[str, dict] = {}

    for cohort in ["ADNI", "AddNeuroMed"]:
        try:
            selected[cohort] = select_full_table(candidates[cohort])
        except RuntimeError as exc:
            available = [
                f"{item['source_file']}::{item['source_sheet']} "
                f"({item['rows']} rows × {item['columns']} columns)"
                for item in candidates[cohort]
            ]

            message = (
                f"\nERROR: {stage} — {cohort}: {exc}\n"
                f"Suitable candidates found: {available or 'None'}"
            )
            raise RuntimeError(message) from exc

    output_path = SUPPLEMENTARY_DIR / config["filename"]

    readme_rows = [
        ["Field", "Value"],
        ["Supplementary table", f"Supplementary Table {config['number']}"],
        ["Title", config["title"]],
        [
            "Analysis policy",
            "Existing analysis outputs were exported without retraining, "
            "model selection, threshold selection, or locked-test access.",
        ],
        [
            "ADNI source",
            f"{selected['ADNI']['source_file']} :: "
            f"{selected['ADNI']['source_sheet']}",
        ],
        [
            "ADNI dimensions",
            f"{selected['ADNI']['rows']} rows × "
            f"{selected['ADNI']['columns']} columns",
        ],
        [
            "AddNeuroMed source",
            f"{selected['AddNeuroMed']['source_file']} :: "
            f"{selected['AddNeuroMed']['source_sheet']}",
        ],
        [
            "AddNeuroMed dimensions",
            f"{selected['AddNeuroMed']['rows']} rows × "
            f"{selected['AddNeuroMed']['columns']} columns",
        ],
        [
            "Selection rule",
            "For each cohort, the largest relevant table in the corresponding "
            "ablation-output directory was selected as the full result table.",
        ],
    ]

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        pd.DataFrame(
            readme_rows[1:],
            columns=readme_rows[0],
        ).to_excel(
            writer,
            sheet_name="README",
            index=False,
        )

        for cohort in ["ADNI", "AddNeuroMed"]:
            item = selected[cohort]
            dataframe = item["dataframe"].copy()

            for sheet_name, part_df in dataframe_parts(dataframe, cohort):
                safe_name = safe_sheet_name(sheet_name)

                part_df.to_excel(
                    writer,
                    sheet_name=safe_name,
                    index=False,
                    startrow=3,
                )

                worksheet = writer.book[safe_name]
                worksheet["A2"] = (
                    f"Cohort: {cohort}; source: "
                    f"{item['source_file']} :: {item['source_sheet']}"
                )

    format_workbook(output_path, config["title"])

    print(f"\nCreated: {output_path}")
    print(
        f"  ADNI: {selected['ADNI']['rows']} rows × "
        f"{selected['ADNI']['columns']} columns"
    )
    print(
        f"    Source: {selected['ADNI']['source_file']} :: "
        f"{selected['ADNI']['source_sheet']}"
    )
    print(
        f"  AddNeuroMed: {selected['AddNeuroMed']['rows']} rows × "
        f"{selected['AddNeuroMed']['columns']} columns"
    )
    print(
        f"    Source: {selected['AddNeuroMed']['source_file']} :: "
        f"{selected['AddNeuroMed']['source_sheet']}"
    )

    return output_path


def main() -> None:
    print("Project root:", ROOT)
    print("Output directory:", SUPPLEMENTARY_DIR)

    created = []

    for stage in ["gene", "mri", "cvae"]:
        created.append(export_stage(stage, TABLE_CONFIG[stage]))

    print("\nAll supplementary ablation tables created successfully:")
    for path in created:
        print(" -", path.relative_to(ROOT))


if __name__ == "__main__":
    main()
